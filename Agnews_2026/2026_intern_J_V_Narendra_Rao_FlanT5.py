"""
Usage:
    python3 run_sheet4_methods.py
    python3 run_sheet4_methods.py --models flan-t5-small --methods Magnitude Movement
"""
import os
import gc
import argparse
import traceback

import torch
import torch.nn as nn
import pandas as pd
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM
from torch.utils.data import DataLoader

from flanutils import (
    BaseConfig, load_data, prepare_dataset_agnews, AGNewsSeq2SeqDataset,
    evaluate_seq2seq, clear_gpu_memory
)

MODEL_CONFIGS = {
    "flan-t5-small": {
        "hf_name": "google/flan-t5-small",
        "batch_size": 32, "grad_accum": 1,
        "max_in": 256, "max_out": 16,
        "calib": 256, "train_n": 20000, "recovery_n": 3000,
        "test_batch": 64,
    },
    "flan-t5-base": {
        "hf_name": "google/flan-t5-base",
        "batch_size": 16, "grad_accum": 1,
        "max_in": 192, "max_out": 16,
        "calib": 192, "train_n": 8000, "recovery_n": 1500,
        "test_batch": 32,
    },
    "flan-t5-large": {
        "hf_name": "google/flan-t5-large",
        "batch_size": 8, "grad_accum": 2,
        "max_in": 128, "max_out": 8,
        "calib": 128, "train_n": 4000, "recovery_n": 1000,
        "test_batch": 16,
    },
    "flan-t5-xl": {
        "hf_name": "google/flan-t5-xl",
        "batch_size": 4, "grad_accum": 4,
        "max_in": 128, "max_out": 8,
        "calib": 96, "train_n": 1000, "recovery_n": 400,
        "test_batch": 8,
    },
}
DISPLAY_NAMES = {
    "flan-t5-small": "Flan-T5-Small", "flan-t5-base": "Flan-T5-Base",
    "flan-t5-large": "Flan-T5-Large", "flan-t5-xl": "Flan-T5-xl",
}
METHODS = ["Magnitude", "Movement", "SparseGPT", "HAWQ", "ZeroQuant"]

PRUNE_RATIO     = 0.20 
LR              = 5e-5
CLIP_NORM       = 1.0
RECOVERY_EPOCHS = 1
FINETUNE_EPOCHS = 1

def clear_gpu():
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()


def model_size_mb(model, path=None):
    if path and os.path.exists(path):
        return os.path.getsize(path) / (1024 ** 2)
    return sum(p.numel() * p.element_size() for p in model.parameters()) / (1024 ** 2)


def get_prunable_params(model):
    """Named weight tensors (dim > 1), excluding lm_head."""
    return [(n, p) for n, p in model.named_parameters()
            if "weight" in n and p.dim() > 1 and "lm_head" not in n]


def get_teacher_checkpoint(model_key, hf_name, train_loader, device):
    artifacts_dir = f"artifacts-{model_key}-agnews"
    os.makedirs(artifacts_dir, exist_ok=True)
    teacher_path = os.path.join(artifacts_dir, "agnews_teacher.pt")

    if os.path.exists(teacher_path):
        print(f"  Found existing checkpoint for {model_key} - reusing")
        return teacher_path

    print(f"  No checkpoint for {model_key} - training from scratch")
    model = AutoModelForSeq2SeqLM.from_pretrained(
        hf_name, torch_dtype=torch.bfloat16,
        device_map="cuda" if torch.cuda.is_available() else "auto",
    )
    grad_accum = MODEL_CONFIGS[model_key]["grad_accum"]

    optimizer = torch.optim.AdamW(model.parameters(), lr=LR)

    for epoch in range(FINETUNE_EPOCHS):
        model.train()
        if hasattr(model, "gradient_checkpointing_enable"):
            model.gradient_checkpointing_enable()
        optimizer.zero_grad()
        for i, batch in enumerate(train_loader):
            ids  = batch["input_ids"].to(device)
            mask = batch["attention_mask"].to(device)
            lbls = batch["labels"].to(device)
            with torch.amp.autocast(device_type="cuda", dtype=torch.bfloat16):
                loss = model(input_ids=ids, attention_mask=mask, labels=lbls).loss / grad_accum
            loss.backward()
            if (i + 1) % grad_accum == 0:
                torch.nn.utils.clip_grad_norm_(model.parameters(), CLIP_NORM)
                optimizer.step()
                optimizer.zero_grad()
            if i % 500 == 0:
                print(f"    [{model_key} ft] epoch {epoch} batch {i} loss {loss.item()*grad_accum:.4f}")

    torch.save(model.state_dict(), teacher_path)
    del model
    clear_gpu()
    return teacher_path


def load_fresh_model(hf_name, teacher_path, device=None):
    model = AutoModelForSeq2SeqLM.from_pretrained(
        hf_name, torch_dtype=torch.bfloat16,
        device_map="cuda" if torch.cuda.is_available() else "auto",
    )
    if teacher_path and os.path.exists(teacher_path):
        model.load_state_dict(torch.load(teacher_path, map_location="cpu"))
        print(f"  Loaded fine-tuned weights from {teacher_path}")
    else:
        print("  No fine-tuned checkpoint - pruning raw pretrained weights")
    if device is not None:
        model = model.to(device)
    return model


def _run_recovery(model, recovery_loader, device, grad_accum):
    lr_rec = 1e-4
    print(f"  Recovery (full FT): {RECOVERY_EPOCHS} epoch(s)  LR={lr_rec}")
    optimizer = torch.optim.AdamW(model.parameters(), lr=lr_rec)

    for ep in range(RECOVERY_EPOCHS):
        model.train()
        if hasattr(model, "gradient_checkpointing_enable"):
            model.gradient_checkpointing_enable()
        optimizer.zero_grad()
        for i, batch in enumerate(recovery_loader):
            ids  = batch["input_ids"].to(device)
            mask = batch["attention_mask"].to(device)
            lbls = batch["labels"].to(device)
            with torch.amp.autocast(device_type="cuda", dtype=torch.bfloat16):
                loss = model(input_ids=ids, attention_mask=mask, labels=lbls).loss / grad_accum
            loss.backward()
            if (i + 1) % grad_accum == 0:
                torch.nn.utils.clip_grad_norm_(model.parameters(), CLIP_NORM)
                optimizer.step()
                optimizer.zero_grad()
            if i % 25 == 0:
                mem = torch.cuda.memory_allocated() / 1024**3 if torch.cuda.is_available() else 0
                print(f"  [recovery ep {ep+1}] batch {i:4d} loss {loss.item()*grad_accum:.4f} GPU {mem:.2f}GB")

    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 1: Magnitude pruning
# ════════════════════════════════════════════════════════════════════════════
def apply_magnitude_pruning(model, prune_ratio=PRUNE_RATIO):
    print(f"  [Magnitude] sparsity={prune_ratio:.0%}")
    with torch.no_grad():
        for name, param in get_prunable_params(model):
            flat = param.data.abs().view(-1)
            k = int(prune_ratio * flat.numel())
            if k == 0:
                continue
            threshold = torch.kthvalue(flat, k).values
            param.data.mul_(param.data.abs() > threshold)
    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 2: Movement pruning — one-shot gradient variant
# ════════════════════════════════════════════════════════════════════════════
def apply_movement_pruning(model, calib_loader, device, prune_ratio=PRUNE_RATIO):
    print(f"  [Movement] collecting gradients (gradient checkpointing enabled) ...")
    model.train()
    if hasattr(model, "gradient_checkpointing_enable"):
        model.gradient_checkpointing_enable()
    model.zero_grad()

    for batch in calib_loader:
        ids  = batch["input_ids"].to(device)
        mask = batch["attention_mask"].to(device)
        lbls = batch["labels"].to(device)
        with torch.amp.autocast(device_type="cuda", dtype=torch.bfloat16):
            loss = model(input_ids=ids, attention_mask=mask, labels=lbls).loss
        loss.backward()
        clear_gpu()

    if hasattr(model, "gradient_checkpointing_disable"):
        model.gradient_checkpointing_disable()

    print(f"  [Movement] applying mask at sparsity={prune_ratio:.0%}")
    with torch.no_grad():
        for name, param in get_prunable_params(model):
            if param.grad is None:
                continue
            scores = (param.data * param.grad).abs()
            k = int(prune_ratio * scores.numel())
            if k == 0:
                param.grad = None
                continue
            threshold = torch.kthvalue(scores.view(-1), k).values
            param.data.mul_(scores > threshold)
            param.grad = None   # free immediately per-tensor

    model.zero_grad()
    clear_gpu()
    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 3: SparseGPT
# ════════════════════════════════════════════════════════════════════════════
class _SparseGPTLayer:
    def __init__(self, layer: nn.Linear):
        self.layer     = layer
        self.n_cols    = layer.weight.shape[1]
        self.H         = torch.zeros(self.n_cols, self.n_cols, dtype=torch.float32)  # on CPU
        self.n_samples = 0

    def add_batch(self, inp):
        x = inp.reshape(-1, self.n_cols).float().cpu()
        self.H        += x.T @ x
        self.n_samples += x.size(0)

    def prune(self, sparsity, device, block_size=128):
        if self.n_samples == 0:
            return
        W = self.layer.weight.data.clone().float()
        H = (self.H / self.n_samples).to(device)
        H.diagonal().add_(0.01 * H.diagonal().mean())   # damping

        try:
            H_inv = torch.cholesky_inverse(torch.linalg.cholesky(H))
        except torch.linalg.LinAlgError:
            H_inv = torch.diag(1.0 / H.diagonal().clamp(min=1e-8))

        mask = torch.zeros_like(W, dtype=torch.bool)
        for start in range(0, self.n_cols, block_size):
            end    = min(start + block_size, self.n_cols)
            W_blk  = W[:, start:end].clone()
            H_blk  = H_inv[start:end, start:end]
            h_diag = H_blk.diagonal().clamp(min=1e-8)

            n_prune = int(sparsity * (end - start))
            if n_prune == 0:
                continue
            scores   = W_blk ** 2 / h_diag.unsqueeze(0)
            thresh   = torch.kthvalue(scores.reshape(-1), n_prune).values
            blk_mask = scores <= thresh
            mask[:, start:end] = blk_mask
            err = (W_blk * blk_mask.float()) / h_diag.unsqueeze(0)
            W[:, start:end] -= err @ H_blk

        W[mask] = 0.0
        self.layer.weight.data = W.to(self.layer.weight.dtype)
        del H, H_inv, W, mask


def apply_sparsegpt(model, calib_loader, device, prune_ratio=PRUNE_RATIO):
    print(f"  [SparseGPT] registering hooks ...")
    sg_layers, hooks = {}, {}

    def make_hook(sg):
        def hook(module, inp, out):
            sg.add_batch(inp[0].detach())
        return hook

    for name, module in model.named_modules():
        if isinstance(module, nn.Linear):
            sg = _SparseGPTLayer(module)
            sg_layers[name] = sg
            hooks[name]     = module.register_forward_hook(make_hook(sg))

    print(f"  [SparseGPT] {len(sg_layers)} Linear layers - calibration pass ...")
    model.eval()
    with torch.no_grad():
        for i, batch in enumerate(calib_loader):
            model(
                input_ids=batch["input_ids"].to(device),
                attention_mask=batch["attention_mask"].to(device),
                labels=batch["labels"].to(device),
            )
            clear_gpu()
            if i % 10 == 0:
                print(f"    calibration batch {i}")

    for h in hooks.values():
        h.remove()

    print(f"  [SparseGPT] pruning at sparsity={prune_ratio:.0%} (layer-by-layer, CPU Hessians) ...")
    for name, sg in sg_layers.items():
        sg.prune(prune_ratio, device)
        clear_gpu()

    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 4: HAWQ
# ════════════════════════════════════════════════════════════════════════════
def apply_hawq_bnb(hf_name):
    from transformers import BitsAndBytesConfig
    import bitsandbytes  # noqa

    print(f"  [bitsandbytes] loading {hf_name} in INT8 ...")
    bnb_config = BitsAndBytesConfig(load_in_8bit=True)
    model = AutoModelForSeq2SeqLM.from_pretrained(
        hf_name,
        quantization_config=bnb_config,
        device_map="auto",
    )
    model.eval()
    print(f"  [bitsandbytes] INT8 model ready.")
    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 5: ZeroQuant 
# ════════════════════════════════════════════════════════════════════════════
def apply_zeroquant_quanto(model, device):
    try:
        from quanto import quantize, freeze, qint8
    except ImportError:
        from optimum.quanto import quantize, freeze, qint8

    print(f"  [quanto] applying INT8 weight quantization ...")
    quantize(model, weights=qint8)
    freeze(model)
    model = model.to(device)
    print(f"  [quanto] quantization complete.")
    return model


# ════════════════════════════════════════════════════════════════════════════
# Harness
# ════════════════════════════════════════════════════════════════════════════

def measure(model, test_loader, tokenizer, device, model_path=None):
    m = evaluate_seq2seq(model, test_loader, tokenizer, device, model_path)
    return {
        "Accuracy (%)": round(m["accuracy"] * 100, 2),
        "Macro-F1(%)":  round(m["f1_score"] * 100, 2),
        "Latency(ms)":  round(m["latency_sec_per_sample"] * 1000, 2),
        "Size (MB)":    round(m["model_size_mb"], 2),
    }


def run_method(
    method_name, model_key, hf_name, teacher_path,
    train_loader, calib_loader, recovery_loader, test_loader,
    tokenizer, device, cfg
):
    print(f"\n{'='*60}")
    print(f"  Method : {method_name.upper()}")
    print(f"  Model  : {model_key}  Sparsity: {PRUNE_RATIO:.0%}")
    print(f"{'='*60}")

    clear_gpu()
    artifacts_dir = f"artifacts-{model_key}-agnews"
    grad_accum = cfg["grad_accum"]

    if method_name == "HAWQ":
        model = apply_hawq_bnb(hf_name)

    elif method_name == "ZeroQuant":
        model = load_fresh_model(hf_name, teacher_path, device=device)
        model = apply_zeroquant_quanto(model, device)

    else:
        model = load_fresh_model(hf_name, teacher_path, device=device)

        if method_name == "Magnitude":
            model = apply_magnitude_pruning(model, PRUNE_RATIO)
        elif method_name == "Movement":
            model = apply_movement_pruning(model, calib_loader, device, PRUNE_RATIO)
        elif method_name == "SparseGPT":
            model = apply_sparsegpt(model, calib_loader, device, PRUNE_RATIO)
        else:
            raise ValueError(f"Unknown method: {method_name}")

        clear_gpu()
        model = _run_recovery(model, recovery_loader, device, grad_accum)

    clear_gpu()

    tmp_path = os.path.join(artifacts_dir, f"{method_name.lower()}_model.pt")
    try:
        torch.save(model.state_dict(), tmp_path)
        print(f"  Saved -> {tmp_path}")
    except Exception as e:
        print(f"  Save skipped ({e})")
        tmp_path = None

    print("  Evaluating ...")
    result = measure(model, test_loader, tokenizer, device, tmp_path)
    del model
    clear_gpu()
    return result


def write_sheet4_format(df, path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet4"
    row_idx = 1
    for method in df["Method"].unique():
        ws.cell(row=row_idx, column=1, value=f"Method: {method}, Dataset: AG News")
        row_idx += 1
        for col, header in enumerate(
            ["Model Name", "Accuracy (%)", "Macro-F1(%)", "Latency(ms)", "Size (MB)"], start=1
        ):
            ws.cell(row=row_idx, column=col, value=header)
        row_idx += 1
        sub = df[df["Method"] == method]
        for _, r in sub.iterrows():
            for col, key in enumerate(
                ["Model Name", "Accuracy (%)", "Macro-F1(%)", "Latency(ms)", "Size (MB)"], start=1
            ):
                ws.cell(row=row_idx, column=col, value=r[key])
            row_idx += 1
        row_idx += 1
    wb.save(path)
    print(f"Wrote Sheet4-formatted results -> {path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--methods", nargs="+", default=METHODS, choices=METHODS)
    parser.add_argument(
        "--models", nargs="+",
        default=list(MODEL_CONFIGS.keys()), choices=list(MODEL_CONFIGS.keys())
    )
    parser.add_argument("--output_csv", default="sheet4_results.csv")
    args = parser.parse_args()

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"Device: {device}")

    train_raw, test_raw = load_data(train_path="agnewstrain.csv", test_path="agnewstest.csv")
    train_df, val_df, test_df = prepare_dataset_agnews(train_raw, test_raw)

    all_rows = []
    for model_key in args.models:
        cfg      = MODEL_CONFIGS[model_key]
        hf_name  = cfg["hf_name"]

        tokenizer = AutoTokenizer.from_pretrained(hf_name)

        calib_df    = train_df.iloc[:cfg["calib"]].reset_index(drop=True)
        recovery_df = train_df.sample(cfg["recovery_n"], random_state=42).reset_index(drop=True)
        base_df     = train_df.sample(cfg["train_n"],    random_state=42).reset_index(drop=True)

        train_loader    = DataLoader(AGNewsSeq2SeqDataset(base_df,     tokenizer), batch_size=cfg["batch_size"], shuffle=True)
        calib_loader    = DataLoader(AGNewsSeq2SeqDataset(calib_df,    tokenizer), batch_size=cfg["batch_size"], shuffle=False)
        recovery_loader = DataLoader(AGNewsSeq2SeqDataset(recovery_df, tokenizer), batch_size=cfg["batch_size"], shuffle=True)
        test_loader     = DataLoader(AGNewsSeq2SeqDataset(test_df,     tokenizer), batch_size=cfg["test_batch"])

        teacher_path = get_teacher_checkpoint(model_key, hf_name, train_loader, device)

        for method_name in args.methods:
            try:
                result = run_method(
                    method_name, model_key, hf_name, teacher_path,
                    train_loader, calib_loader, recovery_loader, test_loader,
                    tokenizer, device, cfg
                )
                all_rows.append({
                    "Method": method_name,
                    "Model Name": DISPLAY_NAMES[model_key],
                    **result,
                })
            except Exception as e:
                print(f"[FAIL] {method_name} on {model_key}: {e}")
                traceback.print_exc()
            finally:
                pd.DataFrame(all_rows).to_csv(args.output_csv, index=False)
                print(f"Saved progress -> {args.output_csv}")
                gc.collect()
                clear_gpu()

    df = pd.DataFrame(all_rows)
    df.to_csv(args.output_csv, index=False)
    if not df.empty:
        write_sheet4_format(df, args.output_csv.replace(".csv", "_sheet4_format.xlsx"))
    print(f"\nAll done.\n{df}")


if __name__ == "__main__":
    main()